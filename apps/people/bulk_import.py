"""Student bulk Excel import — validation + commit phases.

Contract:
- ``parse_workbook`` validates an .xlsx, returns errors per row.
- ``import_rows`` commits in a single transaction; raises if any row fails.

The endpoint runs in two phases:
1. dry_run=True → call ``parse_workbook``, return errors, no DB changes.
2. dry_run=False → call ``parse_workbook`` again, then ``import_rows`` inside
   a transaction. If anything fails we rollback completely — no partial import.

This is currently synchronous (parses ~500 rows in ~1s). When CLAUDE.md
prescribes Celery for bulk imports we'll move ``import_rows`` to a task
and return a job_id; the schema is identical.
"""

from __future__ import annotations

from collections.abc import Iterable
from dataclasses import dataclass, field
from datetime import date, datetime
from io import BytesIO
from typing import Any

from django.db import transaction
from openpyxl import load_workbook

from apps.academics.models import Section, StudentEnrollment
from apps.core.exceptions import ValidationFailed
from apps.people.models import Gender, Student, StudentStatus
from apps.schools.models import School

REQUIRED_HEADERS = [
    "admission_number",
    "first_name",
    "gender",
    "admission_date",
    "class_name",
    "section_name",
]
OPTIONAL_HEADERS = [
    "last_name", "dob", "blood_group", "address",
    "parent1_name", "parent1_phone", "parent1_relation", "parent1_whatsapp",
    "parent2_name", "parent2_phone", "parent2_relation", "parent2_whatsapp",
    "primary_whatsapp_phone", "emergency_contact_name", "emergency_contact_phone",
    "previous_school", "roll_number",
]
ALL_HEADERS = REQUIRED_HEADERS + OPTIONAL_HEADERS


@dataclass
class RowError:
    row: int
    field: str
    message: str


@dataclass
class ParsedRow:
    row: int
    data: dict[str, Any]
    section: Section


@dataclass
class ParseResult:
    rows: list[ParsedRow] = field(default_factory=list)
    errors: list[RowError] = field(default_factory=list)

    @property
    def ok(self) -> bool:
        return not self.errors


def _coerce_date(value: Any) -> date | None:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        d = value.date()
    elif isinstance(value, date):
        d = value
    elif isinstance(value, str):
        d = None
        for fmt in ("%Y-%m-%d", "%d/%m/%Y"):
            try:
                d = datetime.strptime(value, fmt).date()
                break
            except ValueError:
                continue
        if d is None:
            raise ValueError(f"unparseable date '{value}' (use YYYY-MM-DD or DD/MM/YYYY)")
    else:
        raise ValueError(f"unsupported date type {type(value).__name__}")
    today = date.today()
    if not (date(1950, 1, 1) <= d <= date(today.year + 1, 12, 31)):
        raise ValueError(f"date '{d.isoformat()}' is out of range (1950 to next year)")
    return d


def _coerce_bool(value: Any) -> bool:
    if isinstance(value, bool):
        return value
    if value is None or value == "":
        return False
    return str(value).strip().lower() in {"y", "yes", "true", "1"}


def _string(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _select_sheet(wb, name):  # type: ignore[no-untyped-def]
    """Pick the named sheet from a multi-tab workbook, else the active sheet."""
    return wb[name] if name in wb.sheetnames else wb.active


def parse_workbook(*, file_bytes: bytes, school: School) -> ParseResult:
    try:
        wb = load_workbook(BytesIO(file_bytes), data_only=True, read_only=True)
    except Exception as exc:
        raise ValidationFailed("Could not parse Excel file.", {"file": [str(exc)]}) from exc

    ws = _select_sheet(wb, "Students")
    if ws is None:
        raise ValidationFailed("Workbook has no sheets.", {"file": ["empty"]})

    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if not header_row:
        raise ValidationFailed("Sheet is empty.", {"file": ["no rows"]})
    headers = [str(h).strip().lower().replace(" ", "_") if h else "" for h in header_row]

    missing = [h for h in REQUIRED_HEADERS if h not in headers]
    if missing:
        raise ValidationFailed(
            f"Missing required columns: {', '.join(missing)}.",
            {"file": [f"required: {', '.join(REQUIRED_HEADERS)}"]},
        )

    by_idx = {h: i for i, h in enumerate(headers) if h in ALL_HEADERS}

    result = ParseResult()
    seen_admissions: set[str] = set()

    sections_cache: dict[tuple[str, str], Section] = {}

    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if row is None or all(c is None or _string(c) == "" for c in row):
            continue
        errs: list[RowError] = []

        def cell(h: str, *, _row: tuple = row) -> Any:
            if h in by_idx and by_idx[h] < len(_row):
                return _row[by_idx[h]]
            return None

        admission_number = _string(cell("admission_number"))
        first_name = _string(cell("first_name"))
        gender = _string(cell("gender")).title()
        class_name = _string(cell("class_name"))
        section_name = _string(cell("section_name"))

        if not admission_number:
            errs.append(RowError(row_num, "admission_number", "required"))
        elif admission_number in seen_admissions:
            errs.append(RowError(row_num, "admission_number", "duplicate within file"))
        else:
            seen_admissions.add(admission_number)

        if not first_name:
            errs.append(RowError(row_num, "first_name", "required"))
        if gender not in Gender.values:
            errs.append(RowError(row_num, "gender", "must be Male or Female"))

        try:
            admission_date = _coerce_date(cell("admission_date"))
            if admission_date is None:
                errs.append(RowError(row_num, "admission_date", "required"))
        except ValueError as exc:
            errs.append(RowError(row_num, "admission_date", str(exc)))
            admission_date = None

        try:
            dob = _coerce_date(cell("dob"))
        except ValueError as exc:
            errs.append(RowError(row_num, "dob", str(exc)))
            dob = None

        section: Section | None = None
        if class_name and section_name:
            cache_key = (class_name, section_name)
            if cache_key in sections_cache:
                section = sections_cache[cache_key]
            else:
                section = (
                    Section.objects.filter(
                        school=school,
                        class_obj__name=class_name,
                        name=section_name,
                    )
                    .select_related("class_obj__academic_year")
                    .first()
                )
                if section is not None:
                    sections_cache[cache_key] = section
            if section is None:
                errs.append(
                    RowError(row_num, "class_name", f"section '{class_name} {section_name}' not found")
                )
        else:
            if not class_name:
                errs.append(RowError(row_num, "class_name", "required"))
            if not section_name:
                errs.append(RowError(row_num, "section_name", "required"))

        if Student.objects.all_tenants().filter(school=school, admission_number=admission_number).exists():
            errs.append(RowError(row_num, "admission_number", "already exists for this school"))

        if errs:
            result.errors.extend(errs)
            continue

        assert section is not None
        assert admission_date is not None
        result.rows.append(ParsedRow(
            row=row_num,
            data={
                "admission_number": admission_number,
                "first_name": first_name,
                "last_name": _string(cell("last_name")),
                "dob": dob,
                "gender": gender,
                "blood_group": _string(cell("blood_group")),
                "address": _string(cell("address")),
                "admission_date": admission_date,
                "previous_school": _string(cell("previous_school")),
                "primary_whatsapp_phone": _string(cell("primary_whatsapp_phone")),
                "emergency_contact_name": _string(cell("emergency_contact_name")),
                "emergency_contact_phone": _string(cell("emergency_contact_phone")),
                "roll_number": _string(cell("roll_number")),
                "parent1_name": _string(cell("parent1_name")),
                "parent1_phone": _string(cell("parent1_phone")),
                "parent1_relation": _string(cell("parent1_relation")),
                "parent1_whatsapp": _coerce_bool(cell("parent1_whatsapp")),
                "parent2_name": _string(cell("parent2_name")),
                "parent2_phone": _string(cell("parent2_phone")),
                "parent2_relation": _string(cell("parent2_relation")),
                "parent2_whatsapp": _coerce_bool(cell("parent2_whatsapp")),
            },
            section=section,
        ))

    return result


@transaction.atomic
def import_rows(*, school: School, rows: Iterable[ParsedRow]) -> int:
    count = 0
    for parsed in rows:
        d = parsed.data
        student = Student.objects.create(
            school=school,
            admission_number=d["admission_number"],
            first_name=d["first_name"],
            last_name=d["last_name"],
            dob=d["dob"],
            gender=d["gender"],
            blood_group=d["blood_group"],
            address=d["address"],
            admission_date=d["admission_date"],
            previous_school=d["previous_school"],
            primary_whatsapp_phone=d["primary_whatsapp_phone"],
            emergency_contact_name=d["emergency_contact_name"],
            emergency_contact_phone=d["emergency_contact_phone"],
            parent1_name=d["parent1_name"],
            parent1_phone=d["parent1_phone"],
            parent1_relation=d["parent1_relation"],
            parent1_whatsapp=d["parent1_whatsapp"],
            parent2_name=d["parent2_name"],
            parent2_phone=d["parent2_phone"],
            parent2_relation=d["parent2_relation"],
            parent2_whatsapp=d["parent2_whatsapp"],
            status=StudentStatus.ACTIVE,
        )
        StudentEnrollment.objects.create(
            school=school,
            student=student,
            section=parsed.section,
            academic_year=parsed.section.class_obj.academic_year,
            roll_number=d["roll_number"],
            enrollment_date=d["admission_date"],
            status="active",
        )
        count += 1
    return count
