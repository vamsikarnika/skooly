"""Excel export for student lists."""

from __future__ import annotations

from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font

EXPORT_COLUMNS = [
    ("Admission No.", "admission_number"),
    ("First Name", "first_name"),
    ("Last Name", "last_name"),
    ("DOB", "dob"),
    ("Gender", "gender"),
    ("Blood Group", "blood_group"),
    ("Class", None),
    ("Section", None),
    ("Roll", None),
    ("Status", "status"),
    ("Address", "address"),
    ("Parent 1 Name", "parent1_name"),
    ("Parent 1 Phone", "parent1_phone"),
    ("Parent 1 Relation", "parent1_relation"),
    ("Parent 2 Name", "parent2_name"),
    ("Parent 2 Phone", "parent2_phone"),
    ("Admission Date", "admission_date"),
]


def export_students_xlsx(queryset) -> bytes:  # type: ignore[no-untyped-def]
    wb = Workbook()
    ws = wb.active
    ws.title = "Students"

    bold = Font(bold=True)
    for col, (label, _) in enumerate(EXPORT_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col, value=label)
        cell.font = bold

    qs = queryset.prefetch_related("enrollments__section__class_obj")
    for i, student in enumerate(qs, start=2):
        enrollment = next(
            (e for e in student.enrollments.all() if e.status == "active"),
            None,
        )
        for col, (_, field) in enumerate(EXPORT_COLUMNS, start=1):
            if field is None:
                continue
            ws.cell(row=i, column=col, value=getattr(student, field))
        if enrollment:
            ws.cell(row=i, column=7, value=enrollment.section.class_obj.name)
            ws.cell(row=i, column=8, value=enrollment.section.name)
            ws.cell(row=i, column=9, value=enrollment.roll_number)

    for col in ws.columns:
        max_len = max((len(str(c.value)) if c.value is not None else 0) for c in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
