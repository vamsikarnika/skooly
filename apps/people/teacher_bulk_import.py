"""Teacher bulk Excel import — validation + commit phases.

Mirrors apps/people/bulk_import.py (the student importer). Selects the
"Teachers" sheet from a multi-tab workbook, falling back to the active sheet
so single-sheet uploads still work.
"""

from __future__ import annotations

from collections.abc import Iterable
from dataclasses import dataclass, field
from datetime import date, datetime
from io import BytesIO
from typing import Any

from django.db import transaction
from openpyxl import load_workbook

from apps.core.exceptions import ValidationFailed
from apps.people.models import Teacher, TeacherStatus
from apps.schools.models import School

SHEET_NAME = "Teachers"
REQUIRED_HEADERS = ["first_name", "phone"]
OPTIONAL_HEADERS = ["last_name", "email", "qualification", "joining_date"]
ALL_HEADERS = REQUIRED_HEADERS + OPTIONAL_HEADERS


@dataclass
class RowError:
    row: int
    field: str
    message: str


@dataclass
class ParsedRow:
    row: int
    data: dict[str, Any]


@dataclass
class ParseResult:
    rows: list[ParsedRow] = field(default_factory=list)
    errors: list[RowError] = field(default_factory=list)

    @property
    def ok(self) -> bool:
        return not self.errors


def _string(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _coerce_date(value: Any) -> date | None:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        d = value.date()
    elif isinstance(value, date):
        d = value
    elif isinstance(value, str):
        d = None
        for fmt in ("%Y-%m-%d", "%d/%m/%Y"):
            try:
                d = datetime.strptime(value, fmt).date()
                break
            except ValueError:
                continue
        if d is None:
            raise ValueError(f"unparseable date '{value}' (use YYYY-MM-DD or DD/MM/YYYY)")
    else:
        raise ValueError(f"unsupported date type {type(value).__name__}")
    today = date.today()
    if not (date(1950, 1, 1) <= d <= date(today.year + 1, 12, 31)):
        raise ValueError(f"date '{d.isoformat()}' is out of range (1950 to next year)")
    return d


def _select_sheet(wb, name):  # type: ignore[no-untyped-def]
    return wb[name] if name in wb.sheetnames else wb.active


def parse_workbook(*, file_bytes: bytes, school: School) -> ParseResult:
    try:
        wb = load_workbook(BytesIO(file_bytes), data_only=True, read_only=True)
    except Exception as exc:
        raise ValidationFailed("Could not parse Excel file.", {"file": [str(exc)]}) from exc

    ws = _select_sheet(wb, SHEET_NAME)
    if ws is None:
        raise ValidationFailed("Workbook has no sheets.", {"file": ["empty"]})

    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if not header_row:
        raise ValidationFailed("Sheet is empty.", {"file": ["no rows"]})
    headers = [str(h).strip().lower().replace(" ", "_") if h else "" for h in header_row]

    missing = [h for h in REQUIRED_HEADERS if h not in headers]
    if missing:
        raise ValidationFailed(
            f"Missing required columns: {', '.join(missing)}.",
            {"file": [f"required: {', '.join(REQUIRED_HEADERS)}"]},
        )

    by_idx = {h: i for i, h in enumerate(headers) if h in ALL_HEADERS}

    result = ParseResult()
    seen_phones: set[str] = set()

    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if row is None or all(c is None or _string(c) == "" for c in row):
            continue
        errs: list[RowError] = []

        def cell(h: str, *, _row: tuple = row) -> Any:
            if h in by_idx and by_idx[h] < len(_row):
                return _row[by_idx[h]]
            return None

        first_name = _string(cell("first_name"))
        phone = _string(cell("phone"))

        if not first_name:
            errs.append(RowError(row_num, "first_name", "required"))
        if not phone:
            errs.append(RowError(row_num, "phone", "required"))
        elif phone in seen_phones:
            errs.append(RowError(row_num, "phone", "duplicate within file"))
        else:
            seen_phones.add(phone)

        if phone and Teacher.objects.all_tenants().filter(school=school, phone=phone).exists():
            errs.append(RowError(row_num, "phone", "already exists for this school"))

        try:
            joining_date = _coerce_date(cell("joining_date"))
        except ValueError as exc:
            errs.append(RowError(row_num, "joining_date", str(exc)))
            joining_date = None

        if errs:
            result.errors.extend(errs)
            continue

        result.rows.append(
            ParsedRow(
                row=row_num,
                data={
                    "first_name": first_name,
                    "last_name": _string(cell("last_name")),
                    "phone": phone,
                    "email": _string(cell("email")),
                    "qualification": _string(cell("qualification")),
                    "joining_date": joining_date,
                },
            )
        )

    return result


@transaction.atomic
def import_rows(*, school: School, rows: Iterable[ParsedRow]) -> int:
    count = 0
    for parsed in rows:
        d = parsed.data
        Teacher.objects.create(
            school=school,
            user=None,
            first_name=d["first_name"],
            last_name=d["last_name"],
            phone=d["phone"],
            email=d["email"],
            qualification=d["qualification"],
            joining_date=d["joining_date"],
            status=TeacherStatus.ACTIVE,
        )
        count += 1
    return count
