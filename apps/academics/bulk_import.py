"""Class & Section bulk Excel import — validation + commit phases.

Mirrors apps/people/bulk_import.py (the student importer). Selects the
"Classes & Sections" sheet from a multi-tab workbook, falling back to the
active sheet so single-sheet uploads still work. Creates classes (idempotent
per name) and their sections under the school's current academic year.
"""

from __future__ import annotations

from collections.abc import Iterable
from dataclasses import dataclass, field
from io import BytesIO
from typing import Any

from django.db import transaction
from openpyxl import load_workbook

from apps.academics.models import Class, Section
from apps.core.exceptions import ValidationFailed
from apps.schools.models import AcademicYear, School

SHEET_NAME = "Classes & Sections"
REQUIRED_HEADERS = ["class_name", "section_name"]
OPTIONAL_HEADERS = ["room_number"]
ALL_HEADERS = REQUIRED_HEADERS + OPTIONAL_HEADERS


@dataclass
class RowError:
    row: int
    field: str
    message: str


@dataclass
class ParsedRow:
    row: int
    data: dict[str, Any]


@dataclass
class ParseResult:
    rows: list[ParsedRow] = field(default_factory=list)
    errors: list[RowError] = field(default_factory=list)

    @property
    def ok(self) -> bool:
        return not self.errors


def _string(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _select_sheet(wb, name):  # type: ignore[no-untyped-def]
    return wb[name] if name in wb.sheetnames else wb.active


def _current_year(school: School) -> AcademicYear:
    year = AcademicYear.objects.filter(school=school, is_current=True).first()
    if year is None:
        raise ValidationFailed(
            "No current academic year set. Set one in Settings first.",
            {"file": ["no current academic year"]},
        )
    return year


def parse_workbook(*, file_bytes: bytes, school: School) -> ParseResult:
    try:
        wb = load_workbook(BytesIO(file_bytes), data_only=True, read_only=True)
    except Exception as exc:
        raise ValidationFailed("Could not parse Excel file.", {"file": [str(exc)]}) from exc

    ws = _select_sheet(wb, SHEET_NAME)
    if ws is None:
        raise ValidationFailed("Workbook has no sheets.", {"file": ["empty"]})

    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if not header_row:
        raise ValidationFailed("Sheet is empty.", {"file": ["no rows"]})
    headers = [str(h).strip().lower().replace(" ", "_") if h else "" for h in header_row]

    missing = [h for h in REQUIRED_HEADERS if h not in headers]
    if missing:
        raise ValidationFailed(
            f"Missing required columns: {', '.join(missing)}.",
            {"file": [f"required: {', '.join(REQUIRED_HEADERS)}"]},
        )

    year = _current_year(school)
    by_idx = {h: i for i, h in enumerate(headers) if h in ALL_HEADERS}

    result = ParseResult()
    seen: set[tuple[str, str]] = set()

    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if row is None or all(c is None or _string(c) == "" for c in row):
            continue
        errs: list[RowError] = []

        def cell(h: str, *, _row: tuple = row) -> Any:
            if h in by_idx and by_idx[h] < len(_row):
                return _row[by_idx[h]]
            return None

        class_name = _string(cell("class_name"))
        section_name = _string(cell("section_name"))
        room_number = _string(cell("room_number"))

        if not class_name:
            errs.append(RowError(row_num, "class_name", "required"))
        if not section_name:
            errs.append(RowError(row_num, "section_name", "required"))

        if class_name and section_name:
            key = (class_name, section_name)
            if key in seen:
                errs.append(RowError(row_num, "section_name", "duplicate within file"))
            else:
                seen.add(key)
                if Section.objects.filter(
                    school=school,
                    class_obj__name=class_name,
                    class_obj__academic_year=year,
                    name=section_name,
                ).exists():
                    errs.append(
                        RowError(
                            row_num,
                            "section_name",
                            f"section '{class_name} {section_name}' already exists",
                        )
                    )

        if errs:
            result.errors.extend(errs)
            continue

        result.rows.append(
            ParsedRow(
                row=row_num,
                data={
                    "class_name": class_name,
                    "section_name": section_name,
                    "room_number": room_number,
                },
            )
        )

    return result


@transaction.atomic
def import_rows(*, school: School, rows: Iterable[ParsedRow]) -> int:
    year = _current_year(school)
    class_cache: dict[str, Class] = {}
    count = 0
    for parsed in rows:
        d = parsed.data
        cls = class_cache.get(d["class_name"])
        if cls is None:
            cls = Class.objects.filter(
                school=school, academic_year=year, name=d["class_name"]
            ).first()
            if cls is None:
                cls = Class.objects.create(
                    school=school, academic_year=year, name=d["class_name"], display_order=0
                )
            class_cache[d["class_name"]] = cls
        Section.objects.create(
            school=school,
            class_obj=cls,
            name=d["section_name"],
            room_number=d["room_number"],
            capacity=40,
        )
        count += 1
    return count
